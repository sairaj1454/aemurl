import io
import json
import os
from urllib.parse import urlparse, urlunparse

from flask import Flask, render_template_string, request
from openpyxl import load_workbook

IS_VERCEL = bool(os.environ.get("VERCEL"))

app = Flask(__name__)
# Vercel hard-caps request bodies at 4.5 MB; Excel is parsed client-side instead.
app.config["MAX_CONTENT_LENGTH"] = (
    4 * 1024 * 1024 if IS_VERCEL else 10 * 1024 * 1024
)
CACHE_FILE = (
    "/tmp/stats_cache.json"
    if IS_VERCEL
    else os.path.join(os.path.dirname(os.path.abspath(__file__)), "stats_cache.json")
)

# Predefined environments mapping
DOMAINS = {
    "qa": "qa-author.honeywellaerospace.com",
    "stage": "stage-author.honeywellaerospace.com",
    "prod": "author.honeywellaerospace.com",
    "local": "localhost:4502",
    "custom": "",  # Handled dynamically via text input
}


def load_cached_stats():
    """Load the total processed URLs counter from disk cache."""
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r") as f:
                data = json.load(f)
                return data.get("total_processed", 0)
        except Exception:
            return 0
    return 0


def update_cached_stats(count_to_add):
    """Increment and save the total processed counter into disk cache."""
    current_total = load_cached_stats()
    new_total = current_total + count_to_add
    try:
        with open(CACHE_FILE, "w") as f:
            json.dump({"total_processed": new_total}, f)
    except Exception as e:
        print(f"Cache write error: {e}")
    return new_total


HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Advanced AEM URL Transformer</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 40px; max-width: 1000px; color: #333; background-color: #fcfcfc; }
        h2 { color: #222; border-bottom: 2px solid #007bff; padding-bottom: 8px; margin-bottom: 15px; }
        
        /* Dashboard Stats Grid */
        .stats-dashboard { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; }
        .stat-box { background: #fff; border: 1px solid #e2e8f0; border-radius: 6px; padding: 15px; display: flex; flex-direction: column; justify-content: center; box-shadow: 0 2px 4px rgba(0,0,0,0.02); }
        .stat-val { font-size: 24px; font-weight: bold; color: #2b6cb0; margin-bottom: 4px; }
        .stat-label { font-size: 13px; color: #718096; font-weight: 500; }
        .time-saved { color: #2f855a; }
        
        /* New Easy-to-Read Feature Instruction Box */
        .feature-banner { background-color: #f4f9ff; border: 1px solid #bce0ff; border-left: 5px solid #007bff; padding: 20px; border-radius: 6px; margin-bottom: 25px; }
        .feature-banner h3 { margin: 0 0 12px 0; color: #0056b3; font-size: 16px; display: flex; align-items: center; gap: 8px; }
        .steps-container { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px; margin-top: 10px; }
        .step-card { background: #ffffff; padding: 12px; border-radius: 4px; border: 1px solid #dcebfa; font-size: 13px; line-height: 1.4; }
        .step-num { font-weight: bold; color: #007bff; font-size: 14px; margin-bottom: 4px; display: block; }
        
        /* Notice Banner */
        .notice-bar { background-color: #fffaf0; border: 1px solid #feebc8; border-radius: 4px; padding: 10px 15px; margin-bottom: 20px; font-size: 13px; color: #c05621; display: flex; align-items: center; gap: 8px; font-weight: 500; }

        /* Two Column Input Layout */
        .input-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; }
        .input-grid var { font-style: normal; }
        .input-group { display: flex; flex-direction: column; }
        .input-group label { font-weight: bold; margin-bottom: 8px; font-size: 14px; color: #495057; }
        textarea { width: 100%; height: 160px; font-family: monospace; padding: 12px; box-sizing: border-box; border: 1px solid #ced4da; border-radius: 4px; resize: vertical; }
        textarea:focus { border-color: #80bdff; outline: 0; box-shadow: 0 0 0 0.2rem rgba(0,123,255,.25); }
        
        .options-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; margin-bottom: 20px; background: #f1f3f5; padding: 15px; border-radius: 6px; border: 1px solid #e9ecef; }
        .option-group { display: flex; flex-direction: column; }
        .option-group label { font-weight: bold; margin-bottom: 6px; font-size: 14px; }
        select, input[type="text"] { padding: 8px; border: 1px solid #ccc; border-radius: 4px; background-color: white; font-size: 14px; }
        .custom-domain-container { margin-top: 10px; }
        .custom-domain-container input { padding: 8px; width: 100%; font-family: monospace; box-sizing: border-box; }
        .checkbox-group { display: flex; align-items: center; gap: 10px; margin-top: 10px; font-size: 14px; }
        .checkbox-group input { width: 16px; height: 16px; cursor: pointer; }
        
        /* Text Replacement Grid Options Style */
        .replace-pair-grid { display: grid; grid-template-columns: 1fr 1fr; gap: 10px; margin-top: 5px; }
        .replace-pair-grid input { font-family: monospace; width: 100%; box-sizing: border-box; }

        /* Buttons styling */
        .btn-primary { padding: 12px 28px; background-color: #007bff; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 16px; font-weight: 600; transition: background 0.2s; }
        .btn-primary:hover { background-color: #0056b3; }
        .btn-secondary { padding: 6px 12px; background-color: #6c757d; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 13px; min-width: 65px; transition: background 0.2s; }
        .btn-secondary:hover { background-color: #5a6268; }
        .btn-danger { padding: 6px 12px; background-color: #dc3545; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 13px; min-width: 65px; transition: background 0.2s; }
        .btn-danger:hover { background-color: #bd2130; }
        .btn-copy-all { padding: 6px 14px; background-color: #28a745; color: white; border: none; border-radius: 4px; cursor: pointer; font-size: 14px; font-weight: 600; transition: background 0.2s; }
        .btn-copy-all:hover { background-color: #218838; }
        
        .actions-container { display: flex; gap: 6px; align-items: center; }
        .results-controls { display: flex; align-items: center; gap: 20px; }
        
        /* Auto Delete Toggle Switch Layout */
        .toggle-container { display: flex; align-items: center; gap: 8px; font-size: 13px; font-weight: 600; color: #495057; background: #e9ecef; padding: 6px 12px; border-radius: 4px; border: 1px solid #ced4da; }
        .toggle-container input { cursor: pointer; width: 15px; height: 15px; }

        /* Results Display Layout */
        .results-box { margin-top: 35px; border-top: 1px solid #dee2e6; padding-top: 20px; }
        .results-header { display: flex; justify-content: space-between; align-items: center; margin-bottom: 15px; }
        .results-header h3 { margin: 0; color: #333; }
        
        .result-group-block { background: #fdfdfd; border: 1px solid #e2e8f0; border-radius: 6px; padding: 15px; margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.02); transition: all 0.3s ease; }
        .result-group-title { font-size: 13px; font-weight: bold; color: #4a5568; margin-bottom: 10px; border-bottom: 1px dashed #cbd5e0; padding-bottom: 4px; }
        
        .result-item { display: flex; justify-content: space-between; align-items: center; margin-bottom: 8px; padding: 10px 12px; background-color: #ffffff; border: 1px solid #edf2f7; border-radius: 4px; font-family: monospace; font-size: 13px; transition: all 0.2s ease; }
        
        /* Status Tiers */
        .type-original { border-left: 5px solid #007bff; }
        .type-migrated { border-left: 5px solid #28a745; }
        .type-legacy { border-left: 5px solid #ffc107; }
        
        .result-url { word-break: break-all; flex-grow: 1; padding-right: 15px; }
        .result-url a { color: #1a0dab; text-decoration: none; }
        .result-url a:hover { text-decoration: underline; }
        .url-badge { font-size: 11px; font-weight: bold; text-transform: uppercase; padding: 2px 6px; border-radius: 3px; margin-right: 8px; display: inline-block; vertical-align: middle; min-width: 75px; text-align: center; }
        .badge-original { background-color: #ebf8ff; color: #2b6cb0; }
        .badge-migrated { background-color: #f0fff4; color: #22543d; }
        .badge-legacy { background-color: #fffaf0; color: #744210; }
        .badge-unmatched { background-color: #fff5f5; color: #c53030; }
        .badge-matched { background-color: #f0fff4; color: #276749; }
        .badge-ambiguous { background-color: #fffaf0; color: #c05621; }

        .excel-upload-box { background: #f8f9fa; border: 2px dashed #007bff; border-radius: 6px; padding: 20px; margin-bottom: 20px; }
        .excel-upload-box label { font-weight: bold; display: block; margin-bottom: 8px; }
        .excel-upload-box input[type="file"] { font-size: 14px; }
        .excel-hint { font-size: 12px; color: #6c757d; margin-top: 8px; }
        .match-summary { display: flex; gap: 15px; flex-wrap: wrap; margin-bottom: 15px; font-size: 13px; }
        .match-stat { padding: 6px 12px; border-radius: 4px; background: #edf2f7; font-weight: 600; }
        .match-stat.ok { background: #c6f6d5; color: #22543d; }
        .match-stat.warn { background: #feebc8; color: #c05621; }
        .match-stat.err { background: #fed7d7; color: #c53030; }
        .error-bar { background: #fff5f5; border: 1px solid #feb2b2; color: #c53030; padding: 10px 15px; border-radius: 4px; margin-bottom: 15px; font-size: 13px; }
        .slug-label { font-size: 11px; color: #718096; margin-top: 4px; }
        .mode-divider { border-top: 1px solid #dee2e6; margin: 20px 0; text-align: center; color: #6c757d; font-size: 13px; font-weight: 600; }
        .excel-status { font-size: 12px; color: #2b6cb0; margin-top: 8px; font-weight: 600; }
        .btn-primary:disabled { opacity: 0.7; cursor: wait; }
    </style>
    <script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
    <script>
        function toggleCustomInput() {
            var envSelect = document.getElementById("env");
            var customContainer = document.getElementById("custom_domain_container");
            if (envSelect.value === "custom") {
                customContainer.style.display = "block";
                document.getElementById("custom_domain").required = true;
            } else {
                customContainer.style.display = "none";
                document.getElementById("custom_domain").required = false;
            }
        }

        function copyToClipboard(text, button) {
            navigator.clipboard.writeText(text).then(function() {
                var originalText = button.innerText;
                button.innerText = "Copied!";
                button.style.backgroundColor = "#28a745";
                
                var autoDeleteEnabled = document.getElementById("auto_delete_toggle").checked;
                
                if (autoDeleteEnabled) {
                    setTimeout(function() {
                        deleteRow(button);
                    }, 400);
                } else {
                    setTimeout(function() {
                        button.innerText = originalText;
                        button.style.backgroundColor = "";
                    }, 1500);
                }
            }, function(err) {
                console.error('Could not copy text: ', err);
            });
        }

        function deleteRow(button) {
            var row = button.closest('.result-item');
            var parentBlock = button.closest('.result-group-block');
            if (row) { row.remove(); }
            
            if (parentBlock && parentBlock.querySelectorAll('.result-item').length === 0) {
                parentBlock.remove();
            }
            
            var remainingLinks = document.querySelectorAll('.result-url-link');
            if (remainingLinks.length === 0) {
                var resultsBox = document.querySelector('.results-box');
                if (resultsBox) { resultsBox.remove(); }
            }
        }

        function copyAllUrls() {
            var links = document.querySelectorAll('.result-url-link');
            var urls = [];
            links.forEach(function(link) { urls.push(link.href); });
            
            var joinString = urls.join('\\n');
            var copyAllBtn = document.getElementById('copy_all_btn');
            
            navigator.clipboard.writeText(joinString).then(function() {
                var originalText = copyAllBtn.innerText;
                copyAllBtn.innerText = "All Links Copied!";
                copyAllBtn.style.backgroundColor = "#1e7e34";
                setTimeout(function() {
                    copyAllBtn.innerText = originalText;
                    copyAllBtn.style.backgroundColor = "";
                }, 2000);
            });
        }

        function parseExcelRows(rows) {
            var authorCol = null;
            var legacyCol = null;
            var authorLinks = [];
            var legacyPaths = [];

            for (var i = 0; i < rows.length; i++) {
                var row = rows[i] || [];
                if (i === 0) {
                    for (var j = 0; j < row.length; j++) {
                        var lower = String(row[j] || "").toLowerCase();
                        if (lower.indexOf("author") !== -1 && lower.indexOf("link") !== -1) {
                            authorCol = j;
                        } else if (lower.indexOf("legacy") !== -1 && lower.indexOf("path") !== -1) {
                            legacyCol = j;
                        }
                    }
                    if (authorCol === null && legacyCol === null) {
                        authorCol = 0;
                        legacyCol = 1;
                    }
                    continue;
                }

                if (authorCol !== null && row[authorCol]) {
                    authorLinks.push(String(row[authorCol]).trim());
                }
                if (legacyCol !== null && row[legacyCol]) {
                    legacyPaths.push(String(row[legacyCol]).trim());
                }
            }

            return { authorLinks: authorLinks, legacyPaths: legacyPaths };
        }

        function estimatePayloadSize(authorText, legacyText) {
            return new Blob([authorText, legacyText]).size;
        }

        function handleExcelSubmit(event) {
            var form = event.target;
            var fileInput = document.getElementById("excel_file");
            var statusEl = document.getElementById("excel_status");
            var submitBtn = form.querySelector(".btn-primary");

            if (!fileInput || !fileInput.files || !fileInput.files.length) {
                var authorText = document.getElementById("url_input").value;
                var legacyText = document.getElementById("legacy_input").value;
                if (estimatePayloadSize(authorText, legacyText) > 4194304) {
                    event.preventDefault();
                    alert(
                        "Your pasted data is too large for cloud hosting (4.5 MB limit). " +
                        "Split the list into smaller batches and run them separately."
                    );
                }
                return;
            }

            event.preventDefault();

            if (typeof XLSX === "undefined") {
                alert("Excel parser failed to load. Please paste URLs manually or refresh the page.");
                return;
            }

            var file = fileInput.files[0];
            submitBtn.disabled = true;
            submitBtn.value = "Parsing Excel...";
            if (statusEl) {
                statusEl.textContent = "Reading " + file.name + " in your browser (not uploaded to server)...";
            }

            var reader = new FileReader();
            reader.onload = function(loadEvent) {
                try {
                    var workbook = XLSX.read(loadEvent.target.result, { type: "array" });
                    var sheet = workbook.Sheets[workbook.SheetNames[0]];
                    var rows = XLSX.utils.sheet_to_json(sheet, { header: 1, defval: "" });
                    var parsed = parseExcelRows(rows);

                    if (!parsed.authorLinks.length) {
                        throw new Error("No Author Link values found. Check column headers.");
                    }

                    var authorText = parsed.authorLinks.join("\\n");
                    var legacyText = parsed.legacyPaths.join("\\n");
                    if (estimatePayloadSize(authorText, legacyText) > 4194304) {
                        throw new Error(
                            "Extracted data exceeds the 4.5 MB cloud limit. Split the Excel into smaller files."
                        );
                    }

                    document.getElementById("url_input").value = authorText;
                    document.getElementById("legacy_input").value = legacyText;
                    document.getElementById("smart_match").checked = true;
                    fileInput.value = "";

                    if (statusEl) {
                        statusEl.textContent =
                            "Loaded " + parsed.authorLinks.length + " author links and " +
                            parsed.legacyPaths.length + " legacy paths. Transforming...";
                    }

                    submitBtn.disabled = false;
                    submitBtn.value = "Transform & Pair URLs";
                    form.submit();
                } catch (err) {
                    submitBtn.disabled = false;
                    submitBtn.value = "Transform & Pair URLs";
                    if (statusEl) {
                        statusEl.textContent = "";
                    }
                    alert("Excel error: " + err.message);
                }
            };
            reader.onerror = function() {
                submitBtn.disabled = false;
                submitBtn.value = "Transform & Pair URLs";
                if (statusEl) {
                    statusEl.textContent = "";
                }
                alert("Could not read the Excel file.");
            };
            reader.readAsArrayBuffer(file);
        }

        window.onload = function() {
            toggleCustomInput();
            var form = document.querySelector("form");
            if (form) {
                form.addEventListener("submit", handleExcelSubmit);
            }
        };
    </script>
</head>
<body>
    <h2>Advanced AEM URL Transformer</h2>
    
    <div class="stats-dashboard">
        <div class="stat-box">
            <span class="stat-val">{{ total_processed }}</span>
            <span class="stat-label">Total URLs Processed Today</span>
        </div>
        <div class="stat-box">
            <span class="stat-val time-saved">{{ hours_saved }}h {{ mins_saved }}m</span>
            <span class="stat-label">Estimated Manual Entry Time Saved (1.5 min / pair)</span>
        </div>
    </div>

    <div class="notice-bar">
        ⚠️ <strong>Tip:</strong> Avoid refreshing the browser tab unnecessarily to maintain your current visible paired preview workspace lists below.
    </div>
    
    <div class="feature-banner">
        <h3>🚀 Feature Guide: Smart Slug Matching & Excel Upload</h3>
        <div class="steps-container">
            <div class="step-card">
                <span class="step-num">Step 1: Upload Excel or Paste URLs</span>
                Upload an Excel with <strong>Author Link</strong> and <strong>Legacy Path</strong> columns, or paste into the text boxes below.
            </div>
            <div class="step-card">
                <span class="step-num">Step 2: Auto-Match by Slug</span>
                The app extracts the slug from each author link (e.g. <code>pirep-what-expect-when-expecting-fly-procedure</code>) and finds the matching legacy path — even when rows are jumbled.
            </div>
            <div class="step-card">
                <span class="step-num">Step 3: Get Transformed URLs</span>
                Click <strong>"Transform & Pair URLs"</strong> to get migrated author URLs paired with their correct legacy URLs.
            </div>
        </div>
    </div>

    {% if excel_error %}
    <div class="error-bar">{{ excel_error }}</div>
    {% endif %}

    <form method="POST" action="/">

        <div class="excel-upload-box">
            <label for="excel_file">📁 Select Excel (.xlsx) — Author Link + Legacy Path columns</label>
            <input type="file" id="excel_file" accept=".xlsx,.xls">
            <div class="excel-hint">Parsed in your browser — the file is never uploaded. Works on Vercel (4.5 MB cloud limit bypassed). Columns can be jumbled; matching is by page slug.</div>
            <div class="excel-status" id="excel_status"></div>
        </div>

        <div class="mode-divider">— OR paste manually —</div>
        
        <div class="input-grid">
            <div class="input-group">
                <label for="url_input">1. Original URLs / Paths (One per line):</label>
                <textarea name="url_input" id="url_input" placeholder="💡 Paste 1 or multiple paths here...&#10;/content/aerobt/us/en/page1&#10;/content/aerobt/us/en/page2"></textarea>
            </div>
            <div class="input-group">
                <label for="legacy_input">2. Corresponding Legacy URLs / Plain Paths (One per line):</label>
                <textarea name="legacy_input" id="legacy_input" placeholder="💡 Paste matching older paths here...&#10;/content/legacy/us/en/old-page1&#10;/content/legacy/us/en/old-page2"></textarea>
            </div>
        </div>
        
        <div class="options-grid">
            <div class="option-group">
                <label for="env">Target Environment / Domain (For Migrated & Legacy Tiers):</label>
                <select name="env" id="env" onchange="toggleCustomInput()">
                    <option value="qa" {% if selected_opts.env == 'qa' %}selected{% endif %}>QA Author</option>
                    <option value="stage" {% if selected_opts.env == 'stage' %}selected{% endif %}>Stage Author</option>
                    <option value="prod" {% if selected_opts.env == 'prod' %}selected{% endif %}>Prod Author</option>
                    <option value="local" {% if selected_opts.env == 'local' %}selected{% endif %}>Local SDK (localhost:4502)</option>
                    <option value="custom" {% if selected_opts.env == 'custom' %}selected{% endif %}>Use Custom Domain...</option>
                </select>
                
                <div id="custom_domain_container" class="custom-domain-container" style="display: none;">
                    <input type="text" id="custom_domain" name="custom_domain" placeholder="e.g., dev-author.company.com" value="{{ selected_opts.custom_domain }}">
                </div>
            </div>
            
            <div class="option-group">
                <label for="editor_action">Editor Configuration (Migrated URL):</label>
                <select name="editor_action" id="editor_action">
                    <option value="force_add" {% if selected_opts.editor_action == 'force_add' %}selected{% endif %}>Force add /editor.html</option>
                    <option value="keep" {% if selected_opts.editor_action == 'keep' %}selected{% endif %}>Preserve original layout behavior</option>
                    <option value="force_remove" {% if selected_opts.editor_action == 'force_remove' %}selected{% endif %}>Force remove /editor.html</option>
                </select>
            </div>

            <div class="option-group" style="grid-column: span 2;">
                <label>Migrated URL Path Text Replacement (Optional):</label>
                <div class="replace-pair-grid">
                    <input type="text" name="find_text" placeholder="Find segment (e.g., /nda-branding-changes/insights/)" value="{{ selected_opts.find_text }}">
                    <input type="text" name="migrated_replace" placeholder="Replace with (e.g., /uat/insights-old/)" value="{{ selected_opts.migrated_replace }}">
                </div>
            </div>
        </div>

        <div class="checkbox-group">
            <input type="checkbox" id="smart_match" name="smart_match" value="true" {% if selected_opts.smart_match %}checked{% endif %}>
            <label for="smart_match">Smart match by slug (recommended when legacy paths are jumbled / out of order)</label>
        </div>
        <div class="checkbox-group">
            <input type="checkbox" id="use_http" name="use_http" value="true" {% if selected_opts.use_http %}checked{% endif %}>
            <label for="use_http">Force HTTP protocol mapping instead of HTTPS for Migrated & Legacy URLs</label>
        </div>
        <br>
        <input type="submit" class="btn-primary" value="Transform & Pair URLs">
    </form>
    
    {% if detailed_groups %}
    <div class="results-box">
        <div class="results-header">
            <h3>Transformed & Paired Results:</h3>
            <div class="results-controls">
                <div class="toggle-container">
                    <input type="checkbox" id="auto_delete_toggle" name="auto_delete_toggle">
                    <label for="auto_delete_toggle">Auto-Delete on Copy</label>
                </div>
                <button id="copy_all_btn" class="btn-copy-all" onclick="copyAllUrls()">Copy All Generated Links</button>
            </div>
        </div>

        {% if match_summary and match_summary.smart_match %}
        <div class="match-summary">
            <span class="match-stat">Total: {{ match_summary.total }}</span>
            <span class="match-stat ok">Matched: {{ match_summary.matched }}</span>
            {% if match_summary.unmatched > 0 %}
            <span class="match-stat err">Unmatched: {{ match_summary.unmatched }}</span>
            {% endif %}
            {% if match_summary.ambiguous > 0 %}
            <span class="match-stat warn">Ambiguous: {{ match_summary.ambiguous }}</span>
            {% endif %}
        </div>
        {% endif %}
        
        {% for group in detailed_groups %}
            <div class="result-group-block">
                <div class="result-group-title">
                    URL Batch Pair #{{ loop.index }}
                    {% if group.slug %} — slug: <code>{{ group.slug }}</code>{% endif %}
                    {% if group.match_status == 'matched' %}
                    <span class="url-badge badge-matched">Matched</span>
                    {% elif group.match_status == 'unmatched' %}
                    <span class="url-badge badge-unmatched">No Legacy Match</span>
                    {% elif group.match_status == 'ambiguous' %}
                    <span class="url-badge badge-ambiguous">Multiple Matches</span>
                    {% endif %}
                </div>
                
                {% if group.original %}
                <div class="result-item type-original">
                    <div class="result-url">
                        <span class="url-badge badge-original">Original</span>
                        <a href="{{ group.original }}" target="_blank" class="result-url-link">{{ group.original }}</a>
                    </div>
                    <div class="actions-container">
                        <button class="btn-secondary" onclick="copyToClipboard('{{ group.original }}', this)">Copy</button>
                        <button class="btn-danger" onclick="deleteRow(this)">Delete</button>
                    </div>
                </div>
                {% endif %}

                {% if group.migrated %}
                <div class="result-item type-migrated">
                    <div class="result-url">
                        <span class="url-badge badge-migrated">Migrated</span>
                        <a href="{{ group.migrated }}" target="_blank" class="result-url-link">{{ group.migrated }}</a>
                    </div>
                    <div class="actions-container">
                        <button class="btn-secondary" onclick="copyToClipboard('{{ group.migrated }}', this)">Copy</button>
                        <button class="btn-danger" onclick="deleteRow(this)">Delete</button>
                    </div>
                </div>
                {% endif %}

                {% if group.legacy %}
                <div class="result-item type-legacy">
                    <div class="result-url">
                        <span class="url-badge badge-legacy">Legacy</span>
                        <a href="{{ group.legacy }}" target="_blank" class="result-url-link">{{ group.legacy }}</a>
                    </div>
                    <div class="actions-container">
                        <button class="btn-secondary" onclick="copyToClipboard('{{ group.legacy }}', this)">Copy</button>
                        <button class="btn-danger" onclick="deleteRow(this)">Delete</button>
                    </div>
                </div>
                {% endif %}
            </div>
        {% endfor %}
    </div>
    {% endif %}
</body>
</html>
"""


def transform_migrated_url(
    input_str,
    env,
    custom_domain,
    editor_action,
    use_http,
    find_text,
    replace_text,
):
    if not input_str:
        return ""

    input_str = input_str.strip()
    target_domain = (
        custom_domain.replace("http://", "")
        .replace("https://", "")
        .strip("/")
        if (env == "custom" and custom_domain)
        else DOMAINS.get(env, DOMAINS["qa"])
    )
    scheme = "http" if use_http or env == "local" else "https"

    if input_str.startswith("http://") or input_str.startswith("https://"):
        parsed = urlparse(input_str)
        if env == "custom" and not custom_domain:
            target_domain = parsed.netloc
            scheme = parsed.scheme
        path = parsed.path
    else:
        path = input_str if input_str.startswith("/") else f"/{input_str}"

    if find_text and (find_text in path):
        path = path.replace(find_text, replace_text)

    is_editor = path.startswith("/editor.html")
    clean_path = path.replace("/editor.html", "", 1) if is_editor else path

    if editor_action == "force_add":
        path = f"/editor.html{clean_path}"
    elif editor_action == "force_remove":
        path = clean_path
    else:
        path = f"/editor.html{clean_path}" if is_editor else clean_path

    if not path.endswith(".html"):
        path = f"{path}.html"

    return urlunparse((scheme, target_domain, path, "", "", ""))


def format_legacy_url(input_str, env, custom_domain, use_http):
    if not input_str:
        return ""

    input_str = input_str.strip()
    if input_str.startswith("http://") or input_str.startswith("https://"):
        return input_str

    target_domain = (
        custom_domain.replace("http://", "")
        .replace("https://", "")
        .strip("/")
        if (env == "custom" and custom_domain)
        else DOMAINS.get(env, DOMAINS["qa"])
    )
    scheme = "http" if use_http or env == "local" else "https"

    path = input_str if input_str.startswith("/") else f"/{input_str}"

    if not path.endswith(".html"):
        path = f"{path}.html"

    return urlunparse((scheme, target_domain, path, "", "", ""))


def extract_slug(value):
    """Extract the page slug (last path segment, no .html) from a URL or path."""
    if not value:
        return ""

    value = value.strip()
    if value.startswith("http://") or value.startswith("https://"):
        path = urlparse(value).path
    else:
        path = value if value.startswith("/") else f"/{value}"

    if path.startswith("/editor.html"):
        path = path.replace("/editor.html", "", 1)

    segment = path.rstrip("/").split("/")[-1]
    if segment.endswith(".html"):
        segment = segment[:-5]
    return segment.lower()


def build_legacy_lookup(legacy_paths):
    """Map slug -> legacy path. Duplicate slugs are stored as a list."""
    lookup = {}
    for path in legacy_paths:
        slug = extract_slug(path)
        if not slug:
            continue
        lookup.setdefault(slug, []).append(path)
    return lookup


def find_matching_legacy(author_link, legacy_lookup):
    """Find the legacy path whose slug matches the author link slug."""
    slug = extract_slug(author_link)
    if not slug:
        return "", slug, "no_slug"

    matches = legacy_lookup.get(slug, [])
    if len(matches) == 1:
        return matches[0], slug, "matched"
    if len(matches) > 1:
        return matches[0], slug, "ambiguous"
    return "", slug, "unmatched"


def parse_excel_upload(file_storage):
    """Read Author Link and Legacy Path columns from an uploaded Excel file."""
    workbook = load_workbook(
        io.BytesIO(file_storage.read()), read_only=True, data_only=True
    )
    sheet = workbook.active

    author_links = []
    legacy_paths = []
    author_col = None
    legacy_col = None

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if not row:
            continue

        cells = [str(c).strip() if c is not None else "" for c in row]

        if row_idx == 0:
            for col_idx, cell in enumerate(cells):
                lower = cell.lower()
                if "author" in lower and "link" in lower:
                    author_col = col_idx
                elif "legacy" in lower and "path" in lower:
                    legacy_col = col_idx
            if author_col is None and legacy_col is None:
                author_col, legacy_col = 0, 1
            continue

        if author_col is not None and author_col < len(cells) and cells[author_col]:
            author_links.append(cells[author_col])
        if legacy_col is not None and legacy_col < len(cells) and cells[legacy_col]:
            legacy_paths.append(cells[legacy_col])

    workbook.close()
    return author_links, legacy_paths


def build_result_groups(
    author_links,
    legacy_paths,
    env,
    custom_domain,
    editor_action,
    use_http,
    find_text,
    migrated_replace,
    smart_match=False,
):
    """Build paired result groups, optionally matching by slug instead of row order."""
    detailed_groups = []
    legacy_lookup = build_legacy_lookup(legacy_paths) if smart_match else {}

    if smart_match:
        for author_link in author_links:
            raw_legacy, slug, status = find_matching_legacy(author_link, legacy_lookup)
            migrated_url = transform_migrated_url(
                author_link,
                env,
                custom_domain,
                editor_action,
                use_http,
                find_text,
                migrated_replace,
            )
            legacy_url = (
                format_legacy_url(raw_legacy, env, custom_domain, use_http)
                if raw_legacy
                else ""
            )
            detailed_groups.append(
                {
                    "original": author_link,
                    "migrated": migrated_url,
                    "legacy": legacy_url,
                    "slug": slug,
                    "match_status": status,
                }
            )
    else:
        max_length = max(len(author_links), len(legacy_paths))
        for i in range(max_length):
            author_link = author_links[i] if i < len(author_links) else ""
            raw_legacy = legacy_paths[i] if i < len(legacy_paths) else ""
            migrated_url = ""
            if author_link:
                migrated_url = transform_migrated_url(
                    author_link,
                    env,
                    custom_domain,
                    editor_action,
                    use_http,
                    find_text,
                    migrated_replace,
                )
            legacy_url = ""
            if raw_legacy:
                legacy_url = format_legacy_url(
                    raw_legacy, env, custom_domain, use_http
                )
            if author_link or legacy_url:
                detailed_groups.append(
                    {
                        "original": author_link,
                        "migrated": migrated_url,
                        "legacy": legacy_url,
                        "slug": extract_slug(author_link) if author_link else "",
                        "match_status": "row_order",
                    }
                )

    return detailed_groups


def get_form_options():
    return {
        "env": request.form.get("env", "qa"),
        "custom_domain": request.form.get("custom_domain", ""),
        "editor_action": request.form.get("editor_action", "force_add"),
        "use_http": request.form.get("use_http") == "true",
        "find_text": request.form.get("find_text", "").strip(),
        "migrated_replace": request.form.get("migrated_replace", "").strip(),
        "smart_match": request.form.get("smart_match") == "true",
    }


@app.errorhandler(413)
def request_entity_too_large(_error):
    return (
        render_template_string(
            HTML_TEMPLATE,
            detailed_groups=[],
            selected_opts={
                "env": "qa",
                "custom_domain": "",
                "editor_action": "force_add",
                "use_http": False,
                "find_text": "",
                "migrated_replace": "",
                "smart_match": True,
            },
            total_processed=load_cached_stats(),
            hours_saved=0,
            mins_saved=0,
            match_summary=None,
            excel_error=(
                "Request too large for cloud hosting (4.5 MB limit). "
                "Use Excel select (parsed in browser) or split into smaller batches."
            ),
        ),
        413,
    )


@app.route("/", methods=["GET", "POST"])
def home():
    detailed_groups = []
    selected_opts = {
        "env": "qa",
        "custom_domain": "",
        "editor_action": "force_add",
        "use_http": False,
        "find_text": "",
        "migrated_replace": "",
        "smart_match": True,
    }

    # Check current cached counts
    total_processed = load_cached_stats()

    match_summary = None
    excel_error = None

    if request.method == "POST":
        selected_opts = get_form_options()
        author_links = []
        legacy_paths = []

        excel_file = request.files.get("excel_file")
        if excel_file and excel_file.filename:
            try:
                author_links, legacy_paths = parse_excel_upload(excel_file)
                selected_opts["smart_match"] = True
            except Exception as exc:
                excel_error = f"Could not read Excel file: {exc}"
        else:
            raw_input = request.form.get("url_input", "")
            legacy_input = request.form.get("legacy_input", "")
            author_links = [
                line.strip() for line in raw_input.split("\n") if line.strip()
            ]
            legacy_paths = [
                line.strip() for line in legacy_input.split("\n") if line.strip()
            ]

        if not excel_error and author_links:
            detailed_groups = build_result_groups(
                author_links,
                legacy_paths,
                selected_opts["env"],
                selected_opts["custom_domain"],
                selected_opts["editor_action"],
                selected_opts["use_http"],
                selected_opts["find_text"],
                selected_opts["migrated_replace"],
                smart_match=selected_opts["smart_match"],
            )

            matched = sum(
                1 for g in detailed_groups if g.get("match_status") == "matched"
            )
            unmatched = sum(
                1 for g in detailed_groups if g.get("match_status") == "unmatched"
            )
            ambiguous = sum(
                1 for g in detailed_groups if g.get("match_status") == "ambiguous"
            )
            match_summary = {
                "total": len(detailed_groups),
                "matched": matched,
                "unmatched": unmatched,
                "ambiguous": ambiguous,
                "smart_match": selected_opts["smart_match"],
            }

            total_processed = update_cached_stats(len(detailed_groups))

    # Calculate total savings: 1.5 mins per processed pair configuration block
    total_minutes_saved = int(total_processed * 1.5)
    hours_saved = total_minutes_saved // 60
    mins_saved = total_minutes_saved % 60

    return render_template_string(
        HTML_TEMPLATE,
        detailed_groups=detailed_groups,
        selected_opts=selected_opts,
        total_processed=total_processed,
        hours_saved=hours_saved,
        mins_saved=mins_saved,
        match_summary=match_summary,
        excel_error=excel_error,
    )


if __name__ == "__main__":
    app.run(debug=True, port=5000)